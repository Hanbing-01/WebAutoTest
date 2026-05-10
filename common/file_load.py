# !/usr/bin python3
# encoding: utf-8 -*-
# @file     : file_load.py
# @author   : 沙陌 Matongxue_2
# @Time     : 2023/5/14 14:30
# @Copyright: 北京码同学
import openpyxl
import yaml

from paths_manager import buyer_yaml


def read_excel(filepath,sheet_name):
    # 获取整个文档对象
    wb = openpyxl.load_workbook(filepath)
    sheet_data = wb[sheet_name] # 获取某个sheet工作表的数据
    # print(sheet_data)
    lines_count = sheet_data.max_row # 获取总行数
    cols_count = sheet_data.max_column # 获取总列数
    # print(lines_count,cols_count)
    data = [] # 用来存储所有行的数据，每行数据都是一个列表
    # 注意：openpyxl里读取时行号和列号都是从1开始
    for l in range(2,lines_count+1):# l:2,3,4,5,6,7
        line = []  # 用来存储当前行所有的单元格数据
        for c in range(1,cols_count+1):# c:1,2,3,4,5,6
            cell_data = sheet_data.cell(l,c).value
            # print(cell_data)
            if cell_data==None:
                cell_data = ''
            line.append(cell_data)
        data.append(line)
    return data

def load_yaml_file(filepath):
    with open(file=filepath,mode='r',encoding='UTF-8') as f:
        content = yaml.load(f,Loader=yaml.FullLoader)
        return content

def write_yaml(filepath,content):
    with open(file=filepath,mode='w',encoding='UTF-8') as f:
        yaml.dump(content,f,Dumper=yaml.Dumper)

if __name__ == '__main__':
    print(load_yaml_file(buyer_yaml))